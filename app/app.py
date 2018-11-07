# -*- coding: utf-8 -*-
import os
from flask import Flask, request, render_template, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from werkzeug import secure_filename


def generate_keywords(title):
    keys = []
    words_list = title.split(' ')
    for w in range(1, len(words_list)):
        keys.append(' '.join(words_list[w:]))
        keys.append(' '.join(words_list[:-w]))
    return ','.join(keys)    


def create_app():
    app = Flask(__name__)
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

    @app.route('/keywords_generator/', methods=['GET', 'POST'])
    def upload_file():
        if request.method == "POST":
            file_ = request.files['doc']
            if file_:
                wb = load_workbook(file_)
                ws = wb.worksheets[0]
                row_num = 1
                for row in ws:
                    keywords = generate_keywords(row[0].value)
                    ws[f"B{row_num}"] = keywords
                    row_num += 1
                response_file = save_virtual_workbook(wb)
                f = BytesIO()
                f.write(response_file)
                f.seek(0)
                return send_file(f, attachment_filename='keys.xlsx', as_attachment=True)
        return render_template('keywords_generator.html')

    return app
